#!/usr/bin/env python3
"""Build program-data.js from the implementation timeline workbook plus any
newer activity entries logged in data/activity-log.csv.

Reviewable judgment calls (name aliases, Karatu 23 membership, tracked/
not-tracked schools, etc.) live in data/school-taxonomy.yml, not in this
file. Edit that YAML file to change how a school is classified, then re-run
this script.

data/activity-log.csv is the ongoing data-collection intake: new deploy/
upgrade/training/content rows go there (see DATA_COLLECTION_PROCESS.md)
instead of requiring someone to hand-edit the xlsx. Its rows are merged into
the same event list as the xlsx's "By Date" sheet and go through identical
normalization, classification, and validation.

This script also runs a lightweight data-quality check: any two canonical
school names that look like they might be unmerged spelling variants of the
same school (and aren't explicitly confirmed as distinct in the taxonomy)
will fail the build with a report of the suspicious pairs. This is meant to
catch the exact class of bug that name_aliases exists to fix, at build time
instead of silently in the dashboard. Because activity-log.csv rows are
merged before this check runs, a newly logged school with a typo'd name
gets caught the same way a messy timeline row would.
"""

from __future__ import annotations

import csv
import difflib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
TIMELINE = ROOT / "data" / "Implementation_Timeline_for_Website_2007-2025.xlsx"
TAXONOMY_PATH = ROOT / "data" / "school-taxonomy.yml"
ACTIVITY_LOG = ROOT / "data" / "activity-log.csv"
OUT = ROOT / "program-data.js"
CANONICAL_SCHOOLS_OUT = ROOT / "canonical-schools.csv"

ACTIVITY_LOG_REQUIRED_COLUMNS = ["Year", "Month", "Implementation", "School", "District", "Country"]

# Two canonical names with a similarity ratio at or above this threshold are
# flagged as possible unmerged duplicates unless listed in
# confirmed_distinct_pairs.
DUPLICATE_NAME_SIMILARITY_THRESHOLD = 0.82


class TaxonomyError(RuntimeError):
    """Raised when the data-quality checks in the taxonomy fail."""


def load_taxonomy(path: Path = TAXONOMY_PATH) -> dict:
    with path.open() as fh:
        return yaml.safe_load(fh)


def build_name_alias_map(taxonomy: dict) -> dict[str, str]:
    return {key: entry["canonical"] for key, entry in taxonomy["name_aliases"].items()}


def build_meta_label_set(taxonomy: dict) -> set[str]:
    return {item["value"] for item in taxonomy["meta_labels"]}


class Normalizer:
    def __init__(self, taxonomy: dict):
        self.aliases = build_name_alias_map(taxonomy)
        self.meta_labels = build_meta_label_set(taxonomy)
        self.substitutions = taxonomy["text_substitutions"]

    def norm_name(self, raw: str | None) -> str | None:
        if not raw:
            return None
        text = " ".join(str(raw).split())
        key = text.lower()
        if key in self.meta_labels:
            return None
        for sub in self.substitutions:
            if "match" in sub and key == sub["match"]:
                return sub["canonical"]
            if "match_contains" in sub and sub["match_contains"] in key:
                return sub["canonical"]
        return self.aliases.get(key, text)


def slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def classify_activity(label: str) -> str:
    s = label.lower()
    if "shule direct" in s and "train" in s:
        return "training"
    if "shule direct" in s:
        return "content"
    if "upgrade" in s or "replace missing" in s:
        return "upgrade"
    if "train" in s or "workshop" in s:
        return "training"
    if any(token in s for token in ("phase", "sparc", "pi-oneer", "pioneer", "solar", "laptop", "computer", "rachel", "network")):
        return "deploy"
    return "other"


def classify_generation(label: str) -> str | None:
    s = label.lower()
    if "sparc+" in s or "sparc +" in s:
        return "SPARC+"
    if re.search(r"\bsparc\b", s):
        return "SPARC"
    if "pi-oneer" in s or "pioneer" in s:
        return "Pi-oneer"
    if "phase 2" in s:
        return "Phase 2"
    if "phase 1" in s:
        return "Phase 1"
    if "laptop" in s:
        return "Pilot laptop"
    return None


def check_for_unmerged_duplicates(canonical_names: list[str], taxonomy: dict) -> list[tuple[str, str, float]]:
    """Flag canonical name pairs that look like unmerged spelling variants."""
    confirmed_pairs = {
        frozenset(pair["names"]) for pair in taxonomy.get("confirmed_distinct_pairs", [])
    }
    suspects = []
    names = sorted(set(canonical_names))
    for i, a in enumerate(names):
        for b in names[i + 1 :]:
            if frozenset((a, b)) in confirmed_pairs:
                continue
            ratio = difflib.SequenceMatcher(None, a.lower(), b.lower()).ratio()
            if ratio >= DUPLICATE_NAME_SIMILARITY_THRESHOLD:
                suspects.append((a, b, ratio))
    return suspects


class Builder:
    def __init__(self, taxonomy: dict):
        self.taxonomy = taxonomy
        self.normalizer = Normalizer(taxonomy)
        self.original_karatu = taxonomy["original_karatu"]["schools"]
        self.kdp1_named = taxonomy["kdp1_named"]["schools"]
        self.kdp2_named = taxonomy["kdp2_named"]["schools"]
        self.not_tracked = taxonomy["not_tracked"]
        self.computers_by_school = taxonomy["computers_by_school"]
        self.zanzibar_fallback = [
            (row["name"], row["district"], row["region"]) for row in taxonomy["zanzibar_schools_fallback"]
        ]

    def cluster_for(self, school: str, district: str, country: str) -> str:
        if school in self.original_karatu:
            return "Original Karatu"
        if school in self.kdp1_named or school == "Gyekrum Arusha":
            return "KDP 1"
        if school in self.kdp2_named:
            return "KDP 2"
        if school == "Kainam":
            return "Karatu in progress"
        if country == "Peru":
            return "Peru"
        if district and "zanzibar" in district.lower():
            return "Zanzibar"
        mapping = {
            "Bunda": "Bunda",
            "Ngorongoro": "Ngorongoro",
            "Serengeti": "Serengeti",
            "Monduli": "Monduli",
            "Moshi": "Kilimanjaro",
            "Morogoro Urban": "Morogoro",
            "Simanjiro": "Manyara",
            "Arusha": "Arusha other",
        }
        return mapping.get(district, district or "Other")

    def site_type(self, school: str, generations: list[str], activities: list[str]) -> str:
        if school in self.not_tracked and school == "Olturoto":
            return "training_only"
        if school in {"Mgutwa"} or (generations and set(generations) <= {"Pi-oneer"}):
            return "pioneer_only"
        if any(g in {"Phase 1", "Phase 2", "SPARC", "SPARC+", "Pilot laptop"} for g in generations):
            return "full_lab"
        if activities and set(activities) <= {"training"}:
            return "training_only"
        return "unknown"

    def status_for(self, school: str, site: str) -> str:
        if school == "Kainam":
            return "in_progress"
        if school == "Gyekrum Arusha":
            return "proposed"
        if school in self.not_tracked:
            return "inactive"
        if site in {"full_lab", "pioneer_only"}:
            return "active"
        if site == "training_only":
            return "inactive"
        return "unknown"

    def load_activity_log_events(self, path: Path | None = None) -> list[dict]:
        """Parse the ongoing activity-log CSV into the same event shape the
        xlsx loop produces, tagged with source="activity-log" for provenance.

        Returns an empty list if the file doesn't exist or has no data rows
        (e.g. a fresh header-only template) — this is the normal state
        between logged visits, not an error.

        `path` defaults to the module-level ACTIVITY_LOG constant, looked up
        at call time (not baked in as a default parameter value) so tests
        can point it at a temp file by monkeypatching that constant.
        """
        if path is None:
            path = ACTIVITY_LOG
        if not path.exists():
            return []

        with path.open(newline="", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            missing = [c for c in ACTIVITY_LOG_REQUIRED_COLUMNS if c not in (reader.fieldnames or [])]
            if missing:
                raise TaxonomyError(
                    f"{path} is missing required column(s): {missing}. "
                    f"Expected header: {','.join(ACTIVITY_LOG_REQUIRED_COLUMNS)}(,Notes,EnteredBy,EntryDate)"
                )

            events = []
            for row_num, row in enumerate(reader, start=2):  # header is row 1
                year = (row.get("Year") or "").strip()
                impl = (row.get("Implementation") or "").strip()
                school = (row.get("School") or "").strip()
                if not year and not impl and not school:
                    continue  # blank row, e.g. trailing newline
                if not school:
                    raise TaxonomyError(f"{path}:{row_num} has no School value; every logged row needs one.")
                if not impl:
                    raise TaxonomyError(f"{path}:{row_num} ({school}) has no Implementation/activity description.")
                try:
                    year_int = int(year) if year else None
                except ValueError:
                    raise TaxonomyError(f"{path}:{row_num} ({school}) has a non-numeric Year: {year!r}") from None

                label = " ".join(impl.split())
                canonical = self.normalizer.norm_name(school)
                events.append(
                    {
                        "year": year_int,
                        "month": " ".join((row.get("Month") or "").split()),
                        "label": label,
                        "rawSchool": " ".join(school.split()),
                        "school": canonical,
                        "district": " ".join((row.get("District") or "").split()),
                        "country": (row.get("Country") or "").strip() or "Tanzania",
                        "activityType": classify_activity(label),
                        "generation": classify_generation(label),
                        "source": "activity-log",
                        "loggedNote": (row.get("Notes") or "").strip(),
                        "enteredBy": (row.get("EnteredBy") or "").strip(),
                        "entryDate": (row.get("EntryDate") or "").strip(),
                    }
                )
            return events

    def run(self) -> dict:
        wb = load_workbook(TIMELINE, data_only=True)
        by_date = wb["By Date"]
        zanzibar_sheet = wb["Zanzibar Schools"]

        events = []
        for year, month, impl, school, district, country in by_date.iter_rows(min_row=5, max_col=6, values_only=True):
            if not impl and not school:
                continue
            label = " ".join(str(impl).split()) if impl else ""
            canonical = self.normalizer.norm_name(school)
            events.append(
                {
                    "year": int(year) if year else None,
                    "month": str(month).replace("\n", " ").strip() if month else "",
                    "label": label,
                    "rawSchool": " ".join(str(school).split()) if school else "",
                    "school": canonical,
                    "district": " ".join(str(district).split()) if district else "",
                    "country": str(country).strip() if country else "Tanzania",
                    "activityType": classify_activity(label),
                    "generation": classify_generation(label),
                    "source": "timeline",
                }
            )

        activity_log_events = self.load_activity_log_events()
        events.extend(activity_log_events)

        zanzibar_names = []
        for name, region, district in zanzibar_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            if not name:
                continue
            canonical = "Charawe" if str(name).strip() == "7 Districts" else " ".join(str(name).split())
            zanzibar_names.append((canonical, district, region))
        if len(zanzibar_names) < 16:
            zanzibar_names = [(n, d, r) for n, r, d in self.zanzibar_fallback]

        schools: dict[str, dict] = {}

        def ensure(name: str, **kwargs) -> dict:
            row = schools.setdefault(
                name,
                {
                    "schoolId": slug(name),
                    "canonicalName": name,
                    "nameVariants": [],
                    "district": kwargs.get("district", ""),
                    "region": kwargs.get("region", ""),
                    "country": kwargs.get("country", "Tanzania"),
                    "cluster": kwargs.get("cluster", ""),
                    "firstYear": None,
                    "latestYear": None,
                    "currentGeneration": None,
                    "generations": [],
                    "activityCounts": Counter(),
                    "inKaratu23": False,
                    "karatu23Role": None,
                    "notes": [],
                },
            )
            for key in ("district", "region", "country", "cluster"):
                if kwargs.get(key) and not row.get(key):
                    row[key] = kwargs[key]
            return row

        for name, district, region in zanzibar_names:
            ensure(
                name,
                district=str(district or ""),
                region=str(region or "Zanzibar"),
                cluster="Zanzibar",
            )

        for event in events:
            name = event["school"]
            if not name:
                continue
            row = ensure(name, district=event["district"], country=event["country"])
            raw = event["rawSchool"]
            if raw and raw not in row["nameVariants"] and raw != name:
                row["nameVariants"].append(raw)
            if event["year"]:
                row["firstYear"] = event["year"] if row["firstYear"] is None else min(row["firstYear"], event["year"])
                row["latestYear"] = event["year"] if row["latestYear"] is None else max(row["latestYear"], event["year"])
            row["activityCounts"][event["activityType"]] += 1
            if event["generation"] and event["generation"] not in row["generations"]:
                row["generations"].append(event["generation"])
                row["currentGeneration"] = event["generation"]
            if event.get("source") == "activity-log" and event.get("loggedNote"):
                who = f" ({event['enteredBy']})" if event.get("enteredBy") else ""
                when = f" on {event['entryDate']}" if event.get("entryDate") else ""
                row["notes"].append(f"{event['loggedNote']}{who}{when}".strip())

        # Data-quality gate: catch unmerged spelling variants before they
        # silently double-count a school in the dashboard.
        suspects = check_for_unmerged_duplicates(list(schools.keys()), self.taxonomy)
        if suspects:
            lines = [f"  {a!r} <-> {b!r} (similarity {ratio:.2f})" for a, b, ratio in suspects]
            raise TaxonomyError(
                "Possible unmerged school-name duplicates found:\n"
                + "\n".join(lines)
                + "\n\nEither add an alias to data/school-taxonomy.yml (name_aliases) if these "
                "are the same school, or add the pair to confirmed_distinct_pairs if they are "
                "genuinely different schools."
            )

        # Schools known from other archive files but missing from the timeline.
        ensure("Gyekrum Arusha", district="Karatu", country="Tanzania", cluster="KDP 1")["notes"].append(
            "Named in the 2023 Addax & Oryx proposal with Upper Kitete; not on the By Date sheet."
        )
        kainam = ensure("Kainam", district="Karatu", country="Tanzania", cluster="Karatu in progress")
        kainam["notes"].append("From PEF site survey, 18 Feb 2026: solar installed March 2026, training expected May 2026, ~500 students.")
        kainam["currentGeneration"] = "SPARC+"
        kainam["latestYear"] = 2026

        for name, row in schools.items():
            counts = dict(row["activityCounts"])
            activities = [k for k, v in counts.items() if v]
            site = self.site_type(name, row["generations"], activities)
            row["siteType"] = site
            row["status"] = self.status_for(name, site)
            row["cluster"] = row["cluster"] or self.cluster_for(name, row["district"], row["country"])
            row["activityCounts"] = counts
            if name in self.not_tracked:
                row["notes"].append(self.not_tracked[name]["reason"])
                row["status"] = "inactive"
            if row["cluster"] == "Zanzibar":
                row["siteType"] = "pioneer_only"
                row["status"] = "active"
                row["currentGeneration"] = row["currentGeneration"] or "Pi-oneer"
                if "Pi-oneer" not in row["generations"]:
                    row["generations"] = ["Pi-oneer"] + row["generations"]
                row["firstYear"] = row["firstYear"] or 2016
                row["district"] = row["district"] or "Zanzibar"
            if name in self.kdp1_named:
                row["inKaratu23"] = True
                row["karatu23Role"] = "KDP 1 installed"
                row["cluster"] = "KDP 1"
            if name in self.kdp2_named:
                row["inKaratu23"] = True
                row["karatu23Role"] = "KDP 2 installed"
                row["cluster"] = "KDP 2"
            if name == "Gyekrum Arusha":
                row["inKaratu23"] = True
                row["karatu23Role"] = "Proposed; not in timeline"
                row["status"] = "proposed"
            if name == "Kainam":
                row["inKaratu23"] = True
                row["karatu23Role"] = "In progress (site survey)"
                row["status"] = "in_progress"
                row["siteType"] = "full_lab"
            if name in self.original_karatu:
                row["cluster"] = "Original Karatu"
                row["inKaratu23"] = False
                row["karatu23Role"] = "Installed base; not in the 23"

        for name, row in schools.items():
            if row["cluster"] == "Zanzibar":
                row["computers"] = 1
            else:
                row["computers"] = self.computers_by_school.get(name)

        school_list = sorted(schools.values(), key=lambda r: (r["cluster"], r["canonicalName"]))

        named_kdp = [s for s in school_list if s["karatu23Role"] in {"KDP 1 installed", "KDP 2 installed"}]
        karatu23 = {
            "denominator": 23,
            "definition": "Karatu public secondary schools that did not already have a lab when the 2023 district plan was written (32 public schools minus 9 with existing labs). The original 6 PPI Karatu schools are the installed base and are not in this 23.",
            "installedNamed": len(named_kdp),
            "kdp1Named": len(self.kdp1_named),
            "kdp1EquipmentCount": 7,
            "kdp2Named": len(self.kdp2_named),
            "inProgress": 1,
            "proposedUnconfirmed": 1,
            "remainingUnnamed": 23 - 7 - 4,
            "progressForBar": 11,
            "progressLabel": "11 of 23",
            "progressNote": "Uses the equipment file definition: KDP 1 (7) + KDP 2 (4). Timeline names 6 of the 7 KDP 1 schools. Gyekrum Arusha is the likely seventh but is not logged as installed. Kainam is in progress and may occupy one of the remaining 12 slots.",
            "originalBase": self.original_karatu,
            "rows": [],
        }
        for name in self.kdp1_named + ["Gyekrum Arusha"] + self.kdp2_named + ["Kainam"]:
            row = schools[name]
            karatu23["rows"].append(
                {
                    "schoolId": row["schoolId"],
                    "name": name,
                    "status": row["status"],
                    "role": row["karatu23Role"],
                    "year": row["latestYear"],
                    "generation": row["currentGeneration"],
                    "computers": row.get("computers"),
                    "notes": " ".join(row["notes"]),
                }
            )
        karatu23["rows"].append(
            {
                "schoolId": "remaining-unnamed",
                "name": "Remaining Karatu schools not yet in the archive",
                "status": "not_started",
                "role": f"{karatu23['remainingUnnamed']} of 23 still unnamed in these files",
                "year": None,
                "generation": None,
                "computers": None,
                "notes": "Keep as a counted remainder until PEF supplies the official 23-school list.",
            }
        )

        activity_by_year = defaultdict(lambda: Counter())
        for event in events:
            if event["year"]:
                activity_by_year[event["year"]][event["activityType"]] += 1
        years = sorted(activity_by_year)
        activity_series = [
            {
                "year": year,
                "deploy": activity_by_year[year]["deploy"],
                "upgrade": activity_by_year[year]["upgrade"],
                "content": activity_by_year[year]["content"],
                "training": activity_by_year[year]["training"],
                "other": activity_by_year[year]["other"],
            }
            for year in years
        ]

        network = {
            "canonicalSchools": len(school_list),
            "activeFullLabs": sum(1 for s in school_list if s["siteType"] == "full_lab" and s["status"] == "active"),
            "pioneerOnly": sum(1 for s in school_list if s["siteType"] == "pioneer_only" and s["status"] != "inactive"),
            "inactive": sum(1 for s in school_list if s["status"] == "inactive"),
            "inProgress": sum(1 for s in school_list if s["status"] == "in_progress"),
            "proposed": sum(1 for s in school_list if s["status"] == "proposed"),
            "datedTimelineRows": len(events),
            "activityLogRows": sum(1 for e in events if e.get("source") == "activity-log"),
            "trainingRows": sum(1 for e in events if e["activityType"] == "training"),
            "deployRows": sum(1 for e in events if e["activityType"] == "deploy"),
            "knownComputers": sum(s["computers"] or 0 for s in school_list),
        }

        mismatches = [
            {
                "severity": item["severity"],
                "title": item["title"],
                "detail": item["detail"],
            }
            for item in self.taxonomy["known_discrepancies"]
        ]

        known_discrepancies = self.taxonomy["known_discrepancies"]

        metric_contract = [
            {
                "metric": "School served",
                "definition": "Distinct canonical school with a full lab or Pi-oneer. Training-only and group labels do not count.",
            },
            {
                "metric": "Active lab",
                "definition": "Full lab whose latest known status is not inactive. Solar/computer function still needs a current site survey.",
            },
            {
                "metric": "Karatu progress",
                "definition": "KDP installs \u00f7 23. Show the original 6 separately as installed base.",
            },
            {
                "metric": "Students reached",
                "definition": "Sum of latest known enrollment at active schools, labeled as latest headcount \u2014 not 42,000 until a school-year table exists.",
            },
            {
                "metric": "Training",
                "definition": "An activity on an existing site (install week, follow-up, or requested topic). Never +1 school. AJ Poole / Albin Key tab, 2026.",
            },
            {
                "metric": "Alumni percentage",
                "definition": "Named survey + year range + respondent n + exact question. Do not use 10\u00d7 until the comparison group is documented.",
            },
        ]

        payload = {
            "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "sources": {
                "timeline": "Implementation Timeline for Website_2007-2025.xlsx",
                "keyTab": "AJ Poole / Albin meeting notes in the Key tab and PEF Solution Generations Specification",
                "equipment": "PROGRAM EQUIPMENTS_2008-2023.xlsx summaries already extracted into dashboard-data.js",
                "siteSurveys": "PEF School Site Surveys.xlsx (Banjika, Kainam)",
                "proposal": "Powering Potential Proposal UPDATED 2023-06-27.pdf",
            },
            "network": network,
            "karatu23": karatu23,
            "schools": school_list,
            "activityByYear": activity_series,
            "mismatches": mismatches,
            "knownDiscrepancies": known_discrepancies,
            "metricContract": metric_contract,
            "generationKey": [
                {"name": "Phase 1", "detail": "First-gen lab: 5 desktops, open-source software, RACHEL, basic solar."},
                {"name": "Phase 2", "detail": "Same hardware generation, expanded to about 20 computers and a larger solar system."},
                {"name": "Pi-oneer", "detail": "Classroom kit: 1 Raspberry Pi, battery projector, solar charger. Not a full lab."},
                {"name": "SPARC", "detail": "Phase 1 using Raspberry Pi: 5 computers, 3 servers, solar, RACHEL / Office / Shule Direct / Scratch."},
                {"name": "SPARC+", "detail": "Phase 2 using Raspberry Pi: 20 computers and expanded solar. Later years add a projector, then Windows mini PCs."},
                {"name": "Windows mini (unnamed)", "detail": "Newest SPARC+ variant with mini computers and Microsoft Windows/Office. Albin still needs to name this generation."},
            ],
        }
        return payload


def build_payload(taxonomy_path: Path = TAXONOMY_PATH) -> dict:
    taxonomy = load_taxonomy(taxonomy_path)
    return Builder(taxonomy).run()


def write_canonical_schools_csv(school_list: list[dict], path: Path = CANONICAL_SCHOOLS_OUT) -> None:
    """Regenerate the canonical school picklist used for spreadsheet data
    validation (see DATA_COLLECTION_PROCESS.md). Always derived fresh from
    program-data.js's school list so it can never drift from what the
    dashboard actually recognizes.
    """
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["canonicalName", "cluster", "district", "country", "status", "inKaratu23"])
        for school in sorted(school_list, key=lambda s: s["canonicalName"]):
            writer.writerow(
                [
                    school["canonicalName"],
                    school["cluster"],
                    school["district"],
                    school["country"],
                    school["status"],
                    "yes" if school["inKaratu23"] else "no",
                ]
            )


def main() -> None:
    try:
        payload = build_payload()
    except TaxonomyError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    school_list = payload["schools"]
    events_count = payload["network"]["datedTimelineRows"]
    log_count = payload["network"]["activityLogRows"]
    OUT.write_text("window.PROGRAM_DATA = " + json.dumps(payload, indent=2) + ";\n")
    write_canonical_schools_csv(school_list)
    print(f"Wrote {OUT} ({len(school_list)} schools, {events_count} events, {log_count} from activity-log.csv)")
    print(f"Wrote {CANONICAL_SCHOOLS_OUT} ({len(school_list)} rows)")


if __name__ == "__main__":
    main()
